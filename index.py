import os
import re
import subprocess
import time

from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from win10toast import ToastNotifier


class App:
    def __init__(self):
        
        load_dotenv()

        self.notifier = ToastNotifier()

        self.wb = load_workbook("SP.xlsx")
        self.sheet = self.wb.active

        subprocess.Popen([
            os.getenv("CHROME_PATH"),
            f"--remote-debugging-port={os.getenv('DEBUG_PORT')}",
            f"--user-data-dir={os.getenv('USER_DATA_DIR')}"
        ])

        #   Configurações do Chrome para se conectar via DevTools
        options = Options()
        options.debugger_address = f"127.0.0.1:{os.getenv('DEBUG_PORT')}"
        options.add_argument("--start-maximized")
        
        self.navegador = webdriver.Chrome(service=Service(), options=options)
        self.navegador.get(os.getenv("LINK"))

        self.run()

    def logar(self):
        self.navegador.find_element(By.ID, "identificacao").click()

        btnCertificado = WebDriverWait(self.navegador, 10).until(
            EC.presence_of_element_located((By.ID, "linkAbaCertificado"))
        )
        btnCertificado.click()

        time.sleep(3)
        self.navegador.find_element(By.ID, "submitCertificado").click()

    def navegar(self):
        burguerBtn = WebDriverWait(self.navegador, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "header__navbar__menu-hamburger"))
        )
        burguerBtn.click()

        time.sleep(1)
        self.navegador.find_element(By.XPATH, '//*[@id="root"]/div/header/nav/aside[1]/div[1]/nav/ul/li[2]/button').click()
        time.sleep(1)
        self.navegador.find_element(By.XPATH, '//*[@id="root"]/div/header/nav/aside[1]/div[1]/nav/ul/li[2]/ul/li[1]/a').click()

    def pesquisar(self, val):

        print(f"\nPesquisando processo: {val}")
        num_processo = re.sub(r'[^0-9]','', val)

        if num_processo.isdigit() != True:
            print("\n\n❌ Processo não é valida, indo para a proxima.\n\n")
            return False
        
        self.navegador.find_element(By.ID, 'numeroDigitoAnoUnificado').send_keys(num_processo[:13])
        self.navegador.find_element(By.XPATH, "//*[@id='foroNumeroUnificado']").send_keys(num_processo[-4:])
        self.navegador.find_element(By.ID, 'botaoConsultarProcessos').click()
        
    def ponteiro(self):
        for row in self.sheet.iter_rows(min_row=2, max_col=1):
            cell_a = row[0]
            num_processo = str(cell_a.value).strip()
            self.linha = cell_a.row

            yield num_processo

    def polo(self):

        time.sleep(3)
        print("Buscando situação do Polo...")
        elemento = self.navegador.find_element(By.CLASS_NAME, "nomeParteEAdvogado").text
        
        try:
            if os.getenv("NOME_DO_POLO") in elemento.lower():
                print("✅ POLO ATIVO!!!")
            
            else:
                print("❌ POLO INATIVO!!!")
        
        except:
            print("❌ ERRO NA LOCALIZAÇÃO DO POLO!!!")
            raise

    def situProcesso(self):
        print("\nBuscando situação do processo...\n")

        labelSeg = None
        labelSitu = None

        try:
            labelSeg = self.navegador.find_element(By.ID, "labelSegredoDeJusticaProcesso")
            print(f"❌ O processo é um SEGREDO DE JUSTIÇA !!!")
            print("Seguindo para o próximo...")
            return False
        
        except:
            pass

        try:
            labelSitu = self.navegador.find_element(By.ID, "labelSituacaoProcesso")
            situ = labelSitu.text
            print(f"❌ Processo {situ.upper()} !!!")
            print("Seguindo para o próximo...")
            return False
        
        except:
            pass

        if labelSeg is None and labelSitu is None:
            print("✅ NENHUM STATUS ENCONTRADO!!")

    def status(self):

        var = False
        list_status = []

        if "arquivado" in self.navegador.page_source:
            list_status.append("Arquivado")
            print("✅ Caso está ARQUIVADO")
            var = True
            
        if "baixado" in self.navegador.page_source:
            list_status.append("Baixado")
            print("✅ Caso está BAIXADO")
            var = True
            
        if "Julgado Procedente" in self.navegador.page_source:
            list_status.append("Julgado Procedente")
            print("✅ Caso está JULGADO PROCEDENTE")
            var = True

        if "Julgado improcedente" in self.navegador.page_source:
            list_status.append("Julgado Improcedente")
            print("✅ Caso está JULGADO IMPROCEDENTE")
            var = True

        if "sentença" in self.navegador.page_source or "sentenciado" in self.navegador.page_source:
            list_status.append("Sentença")
            print("✅ Caso está SENTENCIADO")
            var = True

        if var == False:
            print("🟨 NENHUM STATUS ENCONTRADO!!\n")

        print() 
        
    def run(self):
        self.logar()
        self.navegar()
        
        for num_processo in self.ponteiro():
            if self.pesquisar(num_processo) != False:
                self.polo()
                if self.situProcesso() != False:
                    self.status()
                
                time.sleep(3)
                print('='*50)
                self.navegador.find_element(By.ID, 'setaVoltar').click()
            
            else:
                print("Seguindo para o próximo processo...")
                continue

try:
    app = App()

except Exception as e:
    print("\n\n❌ ERROR!!!\n\n")